import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="Network Report Auditor", layout="wide")
st.title("📡 Sector & Carrier Keyed Comparison Tool")

# --- STYLES & COLORS ---
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
LIGHT_RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
GREEN_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

# --- SIDEBAR & INTERFACE ---
with st.sidebar:
    st.header("⚙️ Audit Configuration")
    
    # Updated dropdown options to include Top Loaded
    report_name = st.selectbox(
        "Select Report Type",
        options=["Access Distance Histogram", "Abnormal Release", "Cell Footprint", "Top Loaded"],
        key="report_selector"
    )
    
    tech_selection = st.selectbox(
        "Select Technology",
        options=["NR", "LTE", "UMTS", "GSM"],
        key="tech_selector"
    )
    
    st.divider()
    st.header("📋 Audit Instructions")
    st.markdown(f"**Current Mode:** {report_name}")
    st.info("Upload PRE and POST files to begin the comparison.")

# --- FILE UPLOADERS ---
col1, col2 = st.columns(2)
with col1:
    pre_files = st.file_uploader("Upload PRE Reports", accept_multiple_files=True)
with col2:
    post_files = st.file_uploader("Upload POST Reports", accept_multiple_files=True)

# --- HELPER FUNCTIONS ---

def streaming_load(file_obj, sheet_name, p_key, s_key):
    try:
        file_obj.seek(0)
        wb = load_workbook(file_obj, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames: return None
        ws = wb[sheet_name]
        data = []
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > 50: break 
            row_vals = [str(v).strip().lower() if v is not None else "" for v in row]
            
            if s_key:
                found = (p_key.lower() in row_vals and s_key.lower() in row_vals)
            else:
                found = (p_key.lower() in row_vals)

            if found:
                header_row_idx = i
                headers = [str(v).strip() if v is not None else f"Col_{j}" for j, v in enumerate(row)]
                break
        if header_row_idx:
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if any(v is not None for v in row): data.append(row)
            return pd.DataFrame(data, columns=headers)
    except Exception as e:
        st.error(f"Error reading {sheet_name}: {e}")
    return None

def create_comparison_report(df1, df2, p_key, s_key, tech):
    cols_pre_map = {str(c).strip().lower(): c for c in df1.columns}
    cols_post_map = {str(c).strip().lower(): c for c in df2.columns}
    
    actual_p_key = cols_pre_map[p_key.lower()]
    actual_p_post = cols_post_map[p_key.lower()]
    
    if s_key:
        actual_s_key = cols_pre_map[s_key.lower()]
        actual_s_post = cols_post_map[s_key.lower()]
        df1['K'] = df1[actual_p_key].astype(str).str.strip() + '|' + df1[actual_s_key].astype(str).str.strip()
        df2['K'] = df2[actual_p_post].astype(str).str.strip() + '|' + df2[actual_s_post].astype(str).str.strip()
        key_cols = [actual_p_key, actual_s_key]
    else:
        df1['K'] = df1[actual_p_key].astype(str).str.strip()
        df2['K'] = df2[actual_p_post].astype(str).str.strip()
        key_cols = [actual_p_key]

    original_order = [c for c in df1.columns if c not in key_cols + ['K']]
    df1_idx = df1.drop_duplicates(subset='K').set_index('K')
    df2_idx = df2.drop_duplicates(subset='K').set_index('K')
    all_keys = sorted(set(df1_idx.index).union(set(df2_idx.index)))
    
    wb = Workbook()
    ws_det = wb.active
    ws_det.title = "Comparison Results"
    
    headers = key_cols + ['Status']
    for col in original_order:
        headers += [f"{col}\n(PRE)", f"{col}\n(POST)", f"{col}\nMatch?"]
    
    for c_idx, h in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=c_idx, value=h)
        cell.fill, cell.font, cell.border, cell.alignment = HEADER_FILL, HEADER_FONT, THIN_BORDER, CENTER_ALIGN
        ws_det.column_dimensions[cell.column_letter].width = 22

    stats = {"match": 0, "mismatch": 0, "only_pre": 0, "only_post": 0}
    
    for row_idx, key in enumerate(all_keys, 2):
        k_parts = key.split('|') if s_key else [key]
        for i, part in enumerate(k_parts):
            ws_det.cell(row=row_idx, column=i+1, value=part).border = THIN_BORDER
        
        status_col_idx = len(key_cols) + 1
        status_cell = ws_det.cell(row=row_idx, column=status_col_idx)
        status_cell.border = THIN_BORDER

        if key not in df1_idx.index:
            status_cell.value, status_cell.fill, stats["only_post"] = "ONLY IN POST", YELLOW_FILL, stats["only_post"] + 1
        elif key not in df2_idx.index:
            status_cell.value, status_cell.fill, stats["only_pre"] = "ONLY IN PRE", YELLOW_FILL, stats["only_pre"] + 1
        else:
            status_cell.value = "IN BOTH FILES"
            row1, row2 = df1_idx.loc[key], df2_idx.loc[key]
            has_mismatch, col_ptr = False, status_col_idx + 1
            for col in original_order:
                v1, v2 = str(row1.get(col, 'NULL')), str(row2.get(col, 'NULL'))
                c1, c2, m_cell = ws_det.cell(row=row_idx, column=col_ptr, value=v1), ws_det.cell(row=row_idx, column=col_ptr+1, value=v2), ws_det.cell(row=row_idx, column=col_ptr+2)
                if v1 == v2: m_cell.value, m_cell.fill = "✓ MATCH", GREEN_FILL
                else:
                    m_cell.value, m_cell.fill, has_mismatch = "✗ MISMATCH", LIGHT_RED_FILL, True
                    c1.fill, c2.fill = RED_FILL, RED_FILL
                for c in [c1, c2, m_cell]: c.border = THIN_BORDER
                col_ptr += 3
            status_cell.fill, stats["mismatch" if has_mismatch else "match"] = (LIGHT_RED_FILL if has_mismatch else GREEN_FILL), (stats["mismatch" if has_mismatch else "match"] + 1)

    # --- SUMMARY TAB ---
    ws_sum = wb.create_sheet("Summary", 0)
    summary_rows = [
        ["COMPARISON SUMMARY", ""], [""], ["Audit Metadata:", ""],
        ["Report Type", report_name], ["Technology", tech], ["Total Unique Keys", len(all_keys)],
        ["", ""], ["Comparison Results:", ""],
        ["✓ Matching Records", stats["match"]], ["✗ Mismatching Records", stats["mismatch"]],
        ["📄 Records Only in PRE", stats["only_pre"]], ["📄 Records Only in POST", stats["only_post"]],
        ["", ""], ["Legend:", ""]
    ]
    for r_idx, row_data in enumerate(summary_rows, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if r_idx == 1: cell.font, cell.fill = Font(bold=True, size=14, color="FFFFFF"), HEADER_FILL

    legend_data = [("✓ MATCH", GREEN_FILL, "Values match"), ("✗ MISMATCH", LIGHT_RED_FILL, "Mismatch in row"), ("Cell Error", RED_FILL, "Cell level mismatch"), ("Unique", YELLOW_FILL, "Missing record")]
    for i, (text, fill, desc) in enumerate(legend_data):
        c1, c2 = ws_sum.cell(row=len(summary_rows) + i + 1, column=1, value=text), ws_sum.cell(row=len(summary_rows) + i + 1, column=2, value=desc)
        c1.fill, c1.border, c2.border = fill, THIN_BORDER, THIN_BORDER

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# --- MAIN EXECUTION ---
if st.button("🚀 Run Global Audit"):
    if pre_files and post_files:
        pre_dict = {f.name: f for f in pre_files}
        post_dict = {f.name: f for f in post_files}
        zip_buffer = io.BytesIO()
        processed_any = False

        with zipfile.ZipFile(zip_buffer, "w") as zf:
            for fname, fobj in pre_dict.items():
                if fname in post_dict:
                    st.info(f"📁 Processing: {fname}")
                    xl = pd.ExcelFile(fobj, engine='calamine')
                    for i, sname in enumerate(xl.sheet_names):
                        if i == 0 or sname.lower().endswith('pivot') or sname == "General Information":
                            continue
                        
                        # --- MODULAR SWITCH LOGIC ---
                        if report_name == "Top Loaded":
                            if sname == "Sector Summary":
                                primary_key, secondary_key = "Sector Name", None
                            else:
                                primary_key, secondary_key = "Sector Name", "Carrier ID"
                        
                        elif report_name == "Cell Footprint":
                            if sname == "Cell Footprint":
                                primary_key, secondary_key = "Sector Name", None
                            else:
                                primary_key, secondary_key = "Sector Name", "Carrier"
                        
                        else: # Default for Access Distance / Abnormal Release
                            primary_key, secondary_key = "Sector Name", "Carrier"

                        df_pre = streaming_load(fobj, sname, primary_key, secondary_key)
                        df_post = streaming_load(post_dict[fname], sname, primary_key, secondary_key)
                        
                        if df_pre is not None and df_post is not None:
                            st.write(f"⚙️ Analyzing: {sname}...")
                            report_bytes = create_comparison_report(df_pre, df_post, primary_key, secondary_key, tech_selection)
                            zf.writestr(f"{fname.split('.')[0]}/{sname}_Audit.xlsx", report_bytes)
                            processed_any = True
        
        if processed_any:
            st.success(f"🏁 {tech_selection} Audit Complete!")
            st.download_button("📥 Download Results", zip_buffer.getvalue(), "Network_Audit_Results.zip")
        else:
            st.error("No valid data found to compare. Check your Column headers.")
    else:
        st.warning("Please upload both PRE and POST files.")
